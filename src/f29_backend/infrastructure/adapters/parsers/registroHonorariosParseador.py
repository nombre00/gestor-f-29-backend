
import io
import os
import re
import pandas as pd
from bs4 import BeautifulSoup
from dataclasses import dataclass
from typing import Dict, List
from io import BytesIO
from openpyxl import Workbook, load_workbook
import tempfile
# Importación módulos.
from f29_backend.domain.entities.honorarios import RegistroHonorariosMensual

import pandas as pd
# encoding='iso-8859-1'  


def parse_registro_honorarios(bytes_content: bytes) -> RegistroHonorariosMensual:
    # Paso 1: Intentamos leer como HTML desde bytes
    df = None
    try:
        # pd.read_html acepta bytes directamente (desde pandas 1.3+)
        dfs = pd.read_html(io.BytesIO(bytes_content), encoding='iso-8859-1')
        if dfs:
            df = dfs[0]
    except Exception as e:
        print(f"pd.read_html falló: {e}")
    """ try:
        with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp:
            tmp.write(bytes_content)
            tmp_path = tmp.name
        
        dfs = pd.read_html(tmp_path, encoding='iso-8859-1')
        if dfs:
            df = dfs[0]
    except Exception as e:
        print(f"pd.read_html falló: {e}")
    finally:
        os.unlink(tmp_path)  # limpiar el archivo temporal """

    if df is None:
        # Fallback: BeautifulSoup desde bytes
        try:
            soup = BeautifulSoup(bytes_content.decode('latin-1', errors='replace'), 'html.parser')
            table = soup.find('table')
            if table:
                rows = table.find_all('tr')
                data = [[cell.get_text(strip=True) for cell in row.find_all(['td', 'th'])] for row in rows]
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
        except Exception as e:
            raise ValueError(f"No se pudo parsear como HTML: {e} | Filas encontradas: {len(data) if 'data' in locals() else 'N/A'} | Columnas header: {len(data[0]) if 'data' in locals() and data else 'N/A'}")

    if df is None or df.empty:
        raise ValueError("No se encontró tabla válida")

    # Paso 2: Convertir a workbook en memoria (igual que antes)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=True, sheet_name='Honorarios')
    output.seek(0)

    # Paso 3: Cargar el workbook limpio.
    wb = load_workbook(output, read_only=True, data_only=True)
    ws = wb.active

    # Paso 4: Parsear el sheet limpio (ahora con posiciones más confiables).
    contribuyente = ""
    rut = ""
    fecha_informe = ""

    # Iterámos por las filas del encabezado del workbook para capturar los datos.
    for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        row_values = [str(cell or '').strip() for cell in row if cell is not None]
        row_text = ' '.join(row_values)

        # Contribuyente
        if not contribuyente:
            contrib_match = re.search(r'Contribuyente\s*:\s*(.+?)(?:\s*RUT|$)', row_text, re.I)
            if contrib_match:
                contribuyente = contrib_match.group(1).strip()

        # RUT
        if not rut:
            rut_match = re.search(r'RUT\s*:\s*([\d.-]+)', row_text, re.I)
            if rut_match:
                rut = rut_match.group(1).strip()

        # Fecha del informe
        if not fecha_informe:
            informe_match = re.search(r'Informe.*?mes\s*(\d+)\s.*?año\s*(\d+)', row_text, re.I)
            if informe_match:
                mes, ano = informe_match.groups()
                fecha_informe = f"Mes {mes} del año {ano}"
            else:
                # Fallback si no coincide el formato exacto
                if "informe" in row_text.lower() and ("mes" in row_text.lower() or "año" in row_text.lower()):
                    fecha_informe = row_text.strip()

    if not all([contribuyente, rut, fecha_informe]):
        print("DEBUG: Primeras filas del workbook para diagnóstico:")
        for r_idx, r in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            print(f"Fila {r_idx}: {[str(c) for c in r if c is not None]}")
        raise ValueError("No se pudo extraer uno o más campos de cabecera")
    
    #### Hasta acá funciona ####

    # Encontrar fila de encabezados (busca "N°" o "Nro") (Lo encontramos).
    header_idx = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        row_str = ' '.join(str(cell or '').lower() for cell in row if cell is not None)
        if any(term in row_str for term in ['n°', 'nº', 'nro', 'boleta', 'número', 'n° boleta']):
            header_idx = row_idx
            # print(f"DEBUG: Header encontrado en fila {row_idx}")
            break
    if header_idx is None:
        print("DEBUG: No se encontró header. Primeras 30 filas (para ver dónde está N°):")
        for r_idx, r in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            row_str = ' '.join(str(c or '') for c in r if c is not None)
            if 'n' in row_str.lower():  # cualquier cosa con 'n' para debug
                print(f"Fila {r_idx}: {row_str[:100]}...")
        raise ValueError("No se encontró fila de encabezados (N°)")

    # Columnas por posición aproximada.
    col_map = {
        'numero': 0,
        'fecha': 1,
        'estado': 2,
        'rut_emisor': 4,
        'nombre': 5,
        'soc_prof': 6,
        'brutos': 7,
        'retenido': 8,
        'pagado': 9,
    }

    honorarios = []
    totales = {'brutos': 0, 'retenido': 0, 'pagado': 0}

    # Iterámos las boletas.
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):  # Acá nos tenemos que saltar el header encontrado y la siguiente fila que también es un header.
        if not row[0]:
            continue
        first = str(row[0]).lower().strip()
        if 'totales' in first or '(*)' in first:
            totales['brutos'] = int(row[col_map['brutos']] or 0)
            totales['retenido'] = int(row[col_map['retenido']] or 0)
            totales['pagado'] = int(row[col_map['pagado']] or 0)
            break

        boleta = {
            'numero': int(row[col_map['numero']]) if row[col_map['numero']] else None,
            'fecha': str(row[col_map['fecha']]).strip() if row[col_map['fecha']] else '',
            'estado': str(row[col_map['estado']]).strip(),
            'rut_emisor': str(row[col_map['rut_emisor']]).strip(),
            'nombre': str(row[col_map['nombre']]).strip(),
            'soc_prof': str(row[col_map['soc_prof']]).strip(),
            'brutos': int(row[col_map['brutos']] or 0),
            'retenido': int(row[col_map['retenido']] or 0),
            'pagado': int(row[col_map['pagado']] or 0),
        }
        if boleta['numero'] is not None:
            honorarios.append(boleta)

    if totales['brutos'] == 0 and honorarios:
        totales = {
            'brutos': sum(b['brutos'] for b in honorarios),
            'retenido': sum(b['retenido'] for b in honorarios),
            'pagado': sum(b['pagado'] for b in honorarios),
        }

    return RegistroHonorariosMensual(
        contribuyente=contribuyente,
        rut=rut,
        fecha=fecha_informe,
        honorarios=honorarios,
        totales=totales
    )

















""" 
def parse_registro_honorarios(file_path: str) -> RegistroHonorariosMensual:
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    # Paso 1: Leer como HTML (el formato real del SII).
    df = None
    try:
        dfs = pd.read_html(file_path, encoding='iso-8859-1')
        if dfs:
            df = dfs[0]  # asumimos la primera tabla es la principal.
            # print(f"pd.read_html encontró {len(dfs)} tablas, usamos la primera")
    except Exception as e:
        print(f"pd.read_html falló: {e}")

    if df is None:
        # Fallback: leer manualmente con BeautifulSoup.
        try:
            from bs4 import BeautifulSoup
            with open(file_path, encoding='latin-1', errors='replace') as f:
                soup = BeautifulSoup(f, 'html.parser')
            table = soup.find('table')
            if table:
                rows = table.find_all('tr')
                data = [[cell.get_text(strip=True) for cell in row.find_all(['td', 'th'])] for row in rows]
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
                print("Usamos BeautifulSoup para extraer la tabla")
        except ImportError:
            raise ImportError("Instala beautifulsoup4 y lxml: pip install beautifulsoup4 lxml")
        except Exception as e:
            raise ValueError(f"No se pudo leer como HTML: {e}")

    if df is None or df.empty:
        raise ValueError("No se encontró tabla válida en el archivo")

    # Paso 2: Convertir a workbook openpyxl en memoria (limpia el formato).
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=True, sheet_name='Honorarios')
    output.seek(0)

    # Paso 3: Cargar el workbook limpio.
    wb = load_workbook(output, read_only=True, data_only=True)
    ws = wb.active

    # Paso 4: Parsear el sheet limpio (ahora con posiciones más confiables).
    contribuyente = ""
    rut = ""
    fecha_informe = ""

    # Iterámos por las filas del encabezado del workbook para capturar los datos.
    for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        row_values = [str(cell or '').strip() for cell in row if cell is not None]
        row_text = ' '.join(row_values)

        # Contribuyente
        if not contribuyente:
            contrib_match = re.search(r'Contribuyente\s*:\s*(.+?)(?:\s*RUT|$)', row_text, re.I)
            if contrib_match:
                contribuyente = contrib_match.group(1).strip()

        # RUT
        if not rut:
            rut_match = re.search(r'RUT\s*:\s*([\d.-]+)', row_text, re.I)
            if rut_match:
                rut = rut_match.group(1).strip()

        # Fecha del informe
        if not fecha_informe:
            informe_match = re.search(r'Informe.*?mes\s*(\d+)\s.*?año\s*(\d+)', row_text, re.I)
            if informe_match:
                mes, ano = informe_match.groups()
                fecha_informe = f"Mes {mes} del año {ano}"
            else:
                # Fallback si no coincide el formato exacto
                if "informe" in row_text.lower() and ("mes" in row_text.lower() or "año" in row_text.lower()):
                    fecha_informe = row_text.strip()

    if not all([contribuyente, rut, fecha_informe]):
        print("DEBUG: Primeras filas del workbook para diagnóstico:")
        for r_idx, r in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
            print(f"Fila {r_idx}: {[str(c) for c in r if c is not None]}")
        raise ValueError("No se pudo extraer uno o más campos de cabecera")
    
    #### Hasta acá funciona ####

    # Encontrar fila de encabezados (busca "N°" o "Nro") (Lo encontramos).
    header_idx = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        row_str = ' '.join(str(cell or '').lower() for cell in row if cell is not None)
        if any(term in row_str for term in ['n°', 'nº', 'nro', 'boleta', 'número', 'n° boleta']):
            header_idx = row_idx
            # print(f"DEBUG: Header encontrado en fila {row_idx}")
            break
    if header_idx is None:
        print("DEBUG: No se encontró header. Primeras 30 filas (para ver dónde está N°):")
        for r_idx, r in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
            row_str = ' '.join(str(c or '') for c in r if c is not None)
            if 'n' in row_str.lower():  # cualquier cosa con 'n' para debug
                print(f"Fila {r_idx}: {row_str[:100]}...")
        raise ValueError("No se encontró fila de encabezados (N°)")

    # Columnas por posición aproximada.
    col_map = {
        'numero': 0,
        'fecha': 1,
        'estado': 2,
        'rut_emisor': 4,
        'nombre': 5,
        'soc_prof': 6,
        'brutos': 7,
        'retenido': 8,
        'pagado': 9,
    }

    honorarios = []
    totales = {'brutos': 0, 'retenido': 0, 'pagado': 0}

    # Iterámos las boletas.
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):  # Acá nos tenemos que saltar el header encontrado y la siguiente fila que también es un header.
        if not row[0]:
            continue
        first = str(row[0]).lower().strip()
        if 'totales' in first or '(*)' in first:
            totales['brutos'] = int(row[col_map['brutos']] or 0)
            totales['retenido'] = int(row[col_map['retenido']] or 0)
            totales['pagado'] = int(row[col_map['pagado']] or 0)
            break

        boleta = {
            'numero': int(row[col_map['numero']]) if row[col_map['numero']] else None,
            'fecha': str(row[col_map['fecha']]).strip() if row[col_map['fecha']] else '',
            'estado': str(row[col_map['estado']]).strip(),
            'rut_emisor': str(row[col_map['rut_emisor']]).strip(),
            'nombre': str(row[col_map['nombre']]).strip(),
            'soc_prof': str(row[col_map['soc_prof']]).strip(),
            'brutos': int(row[col_map['brutos']] or 0),
            'retenido': int(row[col_map['retenido']] or 0),
            'pagado': int(row[col_map['pagado']] or 0),
        }
        if boleta['numero'] is not None:
            honorarios.append(boleta)

    if totales['brutos'] == 0 and honorarios:
        totales = {
            'brutos': sum(b['brutos'] for b in honorarios),
            'retenido': sum(b['retenido'] for b in honorarios),
            'pagado': sum(b['pagado'] for b in honorarios),
        }

    return RegistroHonorariosMensual(
        contribuyente=contribuyente,
        rut=rut,
        fecha=fecha_informe,
        honorarios=honorarios,
        totales=totales
    ) """
