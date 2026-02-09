
import openpyxl
import pandas as pd
from dataclasses import dataclass
from typing import Dict, List, Optional
from decimal import Decimal
# Importación módulos.
from f29_backend.domain.entities.libroRemuneraciones import LibroRemuneraciones



# Parseador del libro de remuneraciones.
def parse_libro_remuneraciones(file_path: str) -> LibroRemuneraciones:
    wb = openpyxl.load_workbook(file_path, data_only=True)  # Generamos un workbook que contiene los datos del excel.
    
    # ── Hoja 1: Libro Remuneraciones ───────────────────────────────
    sheet_rem = wb['Libro Remuneraciones']  # Guardamos la hoja 1 en una variable. 
    data_rem = [[cell.value for cell in row] for row in sheet_rem.iter_rows()]

    # Función auxiliar.
    def find_row(keyword: str, col: int = 0, data: List[List] = data_rem) -> int | None:
        for i, row in enumerate(data):
            if len(row) > col and keyword.lower() in str(row[col] or '').lower():
                return i
        return None


    # Encabezado (común).
    encabezado = {
        'razon_social': str(data_rem[0][2] or '').strip() if len(data_rem) > 0 else '',
        'rut': str(data_rem[1][2] or '').strip() if len(data_rem) > 1 else '',
        'direccion': str(data_rem[2][2] or '').strip() if len(data_rem) > 2 else '',
        'ciudad': str(data_rem[3][2] or '').strip() if len(data_rem) > 3 else '',
        'periodo': ''
    }
    periodo_idx = find_row('Libro de Remuneraciones mes')
    if periodo_idx is not None:
        texto = str(data_rem[periodo_idx][0] or '').strip()
        if 'mes ' in texto.lower():
            encabezado['periodo'] = texto.split('mes ')[-1].strip()

    # Headers Hoja 1
    header_idx = find_row('ID', data=data_rem) or 7
    headers_rem = [str(cell or '').strip() for cell in data_rem[header_idx]]

    # Filas empleados Hoja 1
    empleados_rem = []
    start_data = header_idx + 2  # salta header + subheader
    for r in range(start_data, len(data_rem)):
        row = data_rem[r]
        if not any(row) or 'total' in str(row[0] or '').lower():
            break
        empleado = {headers_rem[i]: row[i] for i in range(len(headers_rem)) if i < len(row)}
        empleados_rem.append(empleado)

    # ── Hoja 2: Aportes Patronales ───────────────────────────────
    sheet_ap = wb['Aportes Patronales']  # Guardamos la hoja 2 en una variable.
    data_ap = [[cell.value for cell in row] for row in sheet_ap.iter_rows()]

    # Headers Hoja 2
    header_ap_idx = find_row('ID', data=data_ap) or 7
    headers_ap = [str(cell or '').strip() for cell in data_ap[header_ap_idx]]

    # Filas empleados Hoja 2
    empleados_ap = []
    start_ap = header_ap_idx + 1
    for r in range(start_ap, len(data_ap)):
        row = data_ap[r]
        if not any(row) or 'total' in str(row[0] or '').lower():
            break
        empleado_ap = {headers_ap[i]: row[i] for i in range(len(headers_ap)) if i < len(row)}
        empleados_ap.append(empleado_ap)

    # ── Calcular los 3 totales clave ───────────────────────────────
    totales = {
        'total_haberes_remuneraciones': Decimal('0'),
        'impuesto_unico': Decimal('0'),
        'retencion_3porc': Decimal('0')
    }

    # Total Haberes (de Hoja 2)
    idx_haberes_ap = next((i for i, h in enumerate(headers_ap) if 'total haberes' in h.lower()), None)
    if idx_haberes_ap is not None:
        totales['total_haberes_remuneraciones'] = sum(Decimal(str(emp.get(headers_ap[idx_haberes_ap], '0')).replace(',', '.').strip()) for emp in empleados_ap)
        totales['total_haberes_remuneraciones'] = totales['total_haberes_remuneraciones'] /2

    # Impuesto Único (de Hoja 1: Impuesto Renta)
    idx_impuesto = next((i for i, h in enumerate(headers_rem) if 'impuesto renta' in h.lower()), None)
    if idx_impuesto is not None:
        totales['impuesto_unico'] = sum(Decimal(str(emp.get(headers_rem[idx_impuesto], '0')).replace(',', '.').strip()) for emp in empleados_rem)
        totales['impuesto_unico'] = totales['impuesto_unico'] / 2

    # Retención 3% (de Hoja 1: Descuento 3% 21252)
    idx_3porc = next((i for i, h in enumerate(headers_rem) if '3% 21252' in h.lower()), None)
    if idx_3porc is not None:
        totales['retencion_3porc'] = sum(Decimal(str(emp.get(headers_rem[idx_3porc], '0')).replace(',', '.').strip()) for emp in empleados_rem)
        totales['retencion_3porc'] = totales['retencion_3porc'] / 2
    

    return LibroRemuneraciones(
        encabezado=encabezado,
        empleados_remuneraciones=empleados_rem,
        empleados_aportes=empleados_ap, 
        totales=totales
    )
