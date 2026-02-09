
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Cosas por hacer generales:
# Faltan los manejos de errores y excepciones.


def generar_plantilla_resumen_f292() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()  # creamos un archivo excel.
    ws = wb.active  # asignamos la hoja activa por dejecto a una variable.
    ws.title = "resumen F-29"  # renombramos la hoja.


    # por hacer: preguntar que hace cada cosa y probar uno por uno que funciona, algo o todo no funciona.
    # definición de estilos, parece ser un lenguaje de etiquetado similar al html.
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)  # borde definido como angosto.
    thick = Side(style='thick')
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick) # Borde definido como grueso.
    bold = Font(bold=True)
    right = Alignment(horizontal='right')
    left_wrap = Alignment(horizontal='left', wrap_text=True)
    center = Alignment(horizontal='center')
    bottom_center = Alignment(horizontal='center', vertical='bottom')


    # Función helper para poner valor + borde + estilo
    # parámetros: r = fila; c = columna; font = fuente de la letra; alineacion; aplica el borde definido arriba.
    def cell(r, c, value="", font_name="Arial", font_size=10, bold=False, align=None):
        col_letter = openpyxl.utils.get_column_letter(c)  # .get_column_letter() convierte el numero de la columna a letra.
        ref = f"{col_letter}{r}"
        ws[ref] = value
        cell_obj = ws[ref]
        cell_obj.border = border
        cell_obj.font = Font(name=font_name, size=font_size, bold=bold)
        if align:
            cell_obj.alignment = align


    # Función para merge + aplicar borde a todo el rango
    # parámetros: cordenadas que limitan las celdas que unimos y cuyo borde delineamos.
    # por hacer: asignar borde grueso.
    def merge(r_start, c_start, r_end, c_end):
        merge_range = f"{openpyxl.utils.get_column_letter(c_start)}{r_start}:{openpyxl.utils.get_column_letter(c_end)}{r_end}"
        ws.merge_cells(merge_range)
        for r in range(r_start, r_end + 1):
            for c in range(c_start, c_end + 1):
                ws.cell(row=r, column=c).border = border
    

    # Funcion que formatéa los encabezados de cada sección, tienes que revisarlo y corregirlo.
    def agregar_encabezado_columnas(
        fila: int,
        titulos: list[str],
        columna_inicio: int = 4,
        espaciado: int = 2,
    ) -> None:
        thick = Side(style='thick')
        borde_grueso = Border(left=thick, right=thick, top=thick, bottom=thick)
        
        for i, texto in enumerate(titulos):
            col = columna_inicio + (i * espaciado)
            cell(
                r=fila,
                c=col,
                value=texto,
                font=bold,
                align=center,
                border=borde_grueso   # ← aquí el cambio
            )
    

    # Versión acotada de la función, puede resultar más práctica.
    def encabezados_columnas(fila: int, titulos: list[str], col_inicio: int = 4) -> None:
        for i, texto in enumerate(titulos):
            cell(fila, col_inicio + i*2, texto, bold, center)



    # Es pertinente ordenarse por lineas de arriba a abajo.
    # Primero se ingresa el valor, luego se juntan las celdas.
    # ────────────────────────────────────────────────
    # Encabezado (Es desde el 3 al 6 de arriba a abajo y del B a la J horizontalmente).
    # ────────────────────────────────────────────────

    # Linea 3.
    cell(3, 2, "RESUMEN FORMULARIO 29", bold=True, align=center, font_size=12)  # Ingresamos el texto.
    merge(3, 2, 3, 5)  # Juntamos las celdas.
    cell(3, 6, "PERIODO", bold=True, align=center, font_size=12)  # Ingresamos el texto.
    merge(3, 6, 3, 7)  # Juntamos las celdas.
    cell(3, 8, "", bold=True, align=center, font_size=12)  # Ingresamos el texto.
    merge(3, 8, 3, 9)  # Juntamos las celdas.
    cell(3, 10, "", bold=True, align=center, font_size=12)  # Ingresamos el texto.

    # Linea 4.
    cell(4, 2, "CONTRIBUYENTE", bold=True, font_size=11)  # Ingresamos el texto.
    cell(4, 3, "", align=center, font_size=11)
    merge(4, 3, 4, 10)  # Juntamos las celdas.

    # Linea 5.
    cell(5, 2, "RUT", bold=True, font_size=11)   # Ingresamos el texto.
    cell(5, 3, "", align=right, font_size=11)   # Ingresamos el texto.
    merge(5, 3, 5, 5)  # Juntamos las celdas.
    cell(5, 9, "", bold=True, font_size=26, align=center)   # Ingresamos el texto.
    merge(5, 9, 6, 10)  # Juntamos las celdas.
    
    # Linea 6.
    cell(6, 2, "CLAVE SII", bold=True, align=left_wrap, font_size=11)   # Ingresamos el texto.
    cell(6, 3, "", align=right, font_size=11)   # Ingresamos el texto.
    merge(6, 3, 6, 5)  # Juntamos las celdas.
    cell(6, 6, "FOLIO", bold=True, font_size=11)  # Ingresamos el texto.



    # ────────────────────────────────────────────────
    # DÉBITOS Y VENTAS (Es desde 7 al 19 de arriba abajo y de la B a la J horizontalmente).
    # ────────────────────────────────────────────────

    # Linea 7.
    cell(7, 2, "DÉBITOS Y VENTAS", bold=True, align=center)  # Ingresamos el texto.
    merge(7, 2, 7, 10)  # Juntamos las celdas.

    # Linea 8.
    # ws.row_dimensions[8].height = 95  # Cambiamos la altura de esta fila.
    cell(8, 2, align=bottom_center)
    merge(8, 2, 8, 4)  # Juntamos las celdas.
    cell(8, 5, "Monto Neto", align=bottom_center)  # Ingresamos el texto.
    merge(8, 5, 8, 6)  # Juntamos las celdas.
    cell(8, 7, "Cant. Documentos", align=bottom_center)  # Ingresamos el texto.
    merge(8, 7, 8, 8)  # Juntamos las celdas.
    cell(8, 9, "Monto", align=bottom_center)  # Ingresamos el texto.
    merge(8, 9, 8, 10)  # Juntamos las celdas.

    # Linea 9.
    cell(9, 2, "Exportaciones", align=left_wrap)  # Ingresamos el texto.
    merge(9, 2, 9, 4)  # Juntamos las celdas.
    cell(9, 5, "110", align=right)  # Ingresamos el texto.
    cell(9, 6, 0, align=right)  # Ingresamos el texto.
    cell(9, 7, "585", align=right)  # Ingresamos el texto.
    cell(9, 8, 0, align=right)  # Ingresamos el texto.
    cell(9, 9, "20", align=right)  # Ingresamos el texto.
    cell(9, 10, 0, align=right)  # Ingresamos el texto.
    
    # Linea 10.
    cell(10, 2, "Ventas y/o Servicios Exentos o No Gravado", align=left_wrap)  # Ingresamos el texto.
    merge(10, 2, 10, 4)  # Juntamos las celdas.
    cell(10, 5, "34", align=right)  # Ingresamos el texto.
    cell(10, 6, 0, align=right)  # Ingresamos el texto.
    cell(10, 7, "586", align=right)  # Ingresamos el texto.
    cell(10, 8, 0, align=right)  # Ingresamos el texto.
    cell(10, 9, "142", align=right)  # Ingresamos el texto.
    cell(10, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 11.
    cell(11, 2, "Fact.de Compra recibidas con ret.total", align=left_wrap)  # Ingresamos el texto.
    merge(11, 2, 11, 4)  # Juntamos las celdas.
    cell(11, 5, "46", align=right)  # Ingresamos el texto.
    cell(11, 6, 0, align=right)  # Ingresamos el texto.
    cell(11, 7, "515", align=right)  # Ingresamos el texto.
    cell(11, 8, 0, align=right)  # Ingresamos el texto.
    cell(11, 9, "587", align=right)  # Ingresamos el texto.
    cell(11, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 12.
    cell(12, 2, "Facturas emitidas", align=left_wrap)  # Ingresamos el texto.
    merge(12, 2, 12, 4)  # Juntamos las celdas.
    cell(12, 5, "33", align=right)  # Ingresamos el texto.
    cell(12, 6, 0, align=right)  # Ingresamos el texto.
    cell(12, 7, "503", align=right)  # Ingresamos el texto.
    cell(12, 8, 0, align=right)  # Ingresamos el texto.
    cell(12, 9, "502", align=right)  # Ingresamos el texto.
    cell(12, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 13.
    cell(13, 2, "Cant de documentos boletas", align=left_wrap)  # Ingresamos el texto.
    merge(13, 2, 13, 4)  # Juntamos las celdas.
    cell(13, 5, "39", align=right)  # Ingresamos el texto.
    cell(13, 6, 0, align=right)  # Ingresamos el texto.
    cell(13, 7, "110", align=right)  # Ingresamos el texto.
    cell(13, 8, 0, align=right)  # Ingresamos el texto.
    cell(13, 9, "111", align=right)  # Ingresamos el texto.
    cell(13, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 14.
    cell(14, 2, "Comprobantes o Recibos de Trasacciones Medios", align=left_wrap)  # Ingresamos el texto.
    merge(14, 2, 14, 4)  # Juntamos las celdas.
    cell(14, 5, "48", align=right)  # Ingresamos el texto.
    cell(14, 6, 0, align=right)  # Ingresamos el texto.
    cell(14, 7, "758", align=right)  # Ingresamos el texto.
    cell(14, 8, 0, align=right)  # Ingresamos el texto.
    cell(14, 9, "759", align=right)  # Ingresamos el texto.
    cell(14, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 15.
    cell(15, 2, "Notas de Débito emitidas", align=left_wrap)  # Ingresamos el texto.
    merge(15, 2, 15, 4)  # Juntamos las celdas.
    cell(15, 5, "56", align=right)  # Ingresamos el texto.
    cell(15, 6, 0, align=right)  # Ingresamos el texto.
    cell(15, 7, "512", align=right)  # Ingresamos el texto.
    cell(15, 8, 0, align=right)  # Ingresamos el texto.
    cell(15, 9, "513", align=right)  # Ingresamos el texto.
    cell(15, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 16.
    cell(16, 2, "Notas de Crédito emitidas", align=left_wrap)  # Ingresamos el texto.
    merge(16, 2, 16, 4)  # Juntamos las celdas.
    cell(16, 5, "61", align=right)  # Ingresamos el texto.
    cell(16, 6, 0, align=right)  # Ingresamos el texto.
    cell(16, 7, "509", align=right)  # Ingresamos el texto.
    cell(16, 8, 0, align=right)  # Ingresamos el texto.
    cell(16, 9, "510", align=right)  # Ingresamos el texto.
    cell(16, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 17.
    cell(17, 2, "Facturas de Compras recibidas", align=left_wrap)  # Ingresamos el texto.
    merge(17, 2, 17, 4)  # Juntamos las celdas.
    cell(17, 5, "45", align=right)  # Ingresamos el texto.
    cell(17, 6, 0, align=right)  # Ingresamos el texto.
    cell(17, 7, "516", align=right)  # Ingresamos el texto.
    cell(17, 8, 0, align=right)  # Ingresamos el texto.
    cell(17, 9, "517", align=right)  # Ingresamos el texto.
    cell(17, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 18.
    cell(18, 2, "Liquidación y Factura", align=left_wrap)  # Ingresamos el texto.
    merge(18, 2, 18, 4)  # Juntamos las celdas.
    cell(18, 5, "43", align=right)  # Ingresamos el texto.
    cell(18, 6, 0, align=right)  # Ingresamos el texto.
    cell(18, 7, "500", align=right)  # Ingresamos el texto.
    cell(18, 8, 0, align=right)  # Ingresamos el texto.
    cell(18, 9, "501", align=right)  # Ingresamos el texto.
    cell(18, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 19.
    cell(19, 2, "")
    cell(19, 3, "TOTAL NETO", align=center, bold=True)  # Ingresamos el texto.
    merge(19, 3, 19, 5)  # Juntamos las celdas.
    cell(19, 6, 0, align=right, bold=True)  # Ingresamos el texto.
    cell(19, 7, "TOTAL IVA", align=center, bold=True)
    merge(19, 7, 19, 8)  # Juntamos las celdas.
    cell(19, 9, "538", align=center, bold=True)
    cell(19, 10, 0, align=right, bold=True)  # Ingresamos el texto.



    # ────────────────────────────────────────────────
    # CRÉDITOS Y COMPRAS (Es desde 20 al 37 y de la B a la K horizontalmente). 
    # ────────────────────────────────────────────────

    # Linea 20.
    cell(20, 2, "CRÉDITOS Y COMPRAS", align=center, bold=True, font_size=11)
    merge(20, 2, 20, 10)

    # Linea 21.
    merge(21, 2, 21, 4)  # Juntamos las celdas.
    cell(21, 5, "Con derecho", align=center)  # Ingresamos el texto.
    merge(21, 5, 21, 6)  # Juntamos las celdas.
    cell(21, 7, "Sin derecho a Crédito", align=center)  # Ingresamos el texto.
    merge(21, 7, 21, 10)  # Juntamos las celdas.

    # Linea 22.
    cell(22, 2, "IVA Por Facturas Electrónicas", align=left_wrap)  # Ingresamos el texto.
    merge(22, 2, 22, 4)  # Juntamos las celdas.
    cell(22, 5, "511", align=right)  # Ingresamos el texto.
    cell(22, 6, 0, align=right)  # Ingresamos el texto.
    cell(22, 7, "514", align=right)  # Ingresamos el texto.
    cell(22, 8, 0, align=right)  # Ingresamos el texto.
    merge(22, 8, 22, 10)  # Juntamos las celdas.

    # Linea 23.
    merge(23, 2, 23, 4)  # Juntamos las celdas.
    cell(23, 5, "Monto Neto", align=center)  # Ingresamos el texto.
    merge(23, 5, 23, 6)  # Juntamos las celdas.
    cell(23, 7, "Cant. Documentos", align=center)  # Ingresamos el texto.
    merge(23, 7, 23, 8)  # Juntamos las celdas.
    cell(23, 9, "Monto", align=center)  # Ingresamos el texto.
    merge(23, 9, 23, 10)  # Juntamos las celdas.

    # Linea 24.
    cell(24, 2, "Compras Internas Afectadas (Sin derecho a Cdto.Fiscal)", align=left_wrap)  # Ingresamos el texto.
    merge(24, 2, 24, 4)  # Juntamos las celdas.
    cell(24, 5, "", align=right)  # Ingresamos el texto.
    cell(24, 6, 0, align=right)  # Ingresamos el texto.
    cell(24, 7, "564", align=right)  # Ingresamos el texto.
    cell(24, 8, 0, align=right)  # Ingresamos el texto.
    cell(24, 9, "521", align=right)  # Ingresamos el texto.
    cell(24, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 25.
    cell(25, 2, "Importaciones (Sin derecho a Cdto.Fiscal)", align=left_wrap)  # Ingresamos el texto.
    merge(25, 2, 25, 4)  # Juntamos las celdas.
    cell(25, 5, "", align=right)  # Ingresamos el texto.
    cell(25, 6, 0, align=right)  # Ingresamos el texto.
    cell(25, 7, "566", align=right)  # Ingresamos el texto.
    cell(25, 8, 0, align=right)  # Ingresamos el texto.
    cell(25, 9, "560", align=right)  # Ingresamos el texto.
    cell(25, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 26.
    cell(26, 2, "Compras Internas exentas o No gravadas", align=left_wrap)  # Ingresamos el texto.
    merge(26, 2, 26, 4)  # Juntamos las celdas.
    cell(26, 5, "34", align=right)  # Ingresamos el texto.
    cell(26, 6, 0, align=right)  # Ingresamos el texto.
    cell(26, 7, "584", align=right)  # Ingresamos el texto.
    cell(26, 8, 0, align=right)  # Ingresamos el texto.
    cell(26, 9, "562", align=right)  # Ingresamos el texto.
    cell(26, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 27.
    merge(27, 2, 27, 4)  # Juntamos las celdas.
    cell(27, 5, "Monto Neto", align=center)  # Ingresamos el texto.
    merge(27, 5, 27, 6)  # Juntamos las celdas.
    cell(27, 7, "Cant. Documentos", align=center)  # Ingresamos el texto.
    merge(27, 7, 27, 8)  # Juntamos las celdas.
    cell(27, 9, "Monto", align=center)  # Ingresamos el texto.
    merge(27, 9, 27, 10)  # Juntamos las celdas.

    # Linea 28.
    cell(28, 2, "Facturas recibidas y Facturas de compra", align=left_wrap)  # Ingresamos el texto.
    merge(28, 2, 28, 4)  # Juntamos las celdas.
    cell(28, 5, "33", align=right)  # Ingresamos el texto.
    cell(28, 6, 0, align=right)  # Ingresamos el texto.
    cell(28, 7, "519", align=right)  # Ingresamos el texto.
    cell(28, 8, 0, align=right)  # Ingresamos el texto.
    cell(28, 9, "520", align=right)  # Ingresamos el texto.
    cell(28, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 29.
    cell(29, 2, "Fact.Recib.Prov.Superm.Art.33 N°4 DL825(Ley 20,780)", align=left_wrap)  # Ingresamos el texto.
    merge(29, 2, 29, 4)  # Juntamos las celdas.
    cell(29, 5, "", align=right)  # Ingresamos el texto.
    cell(29, 6, 0, align=right)  # Ingresamos el texto.
    cell(29, 7, "761", align=right)  # Ingresamos el texto.
    cell(29, 8, 0, align=right)  # Ingresamos el texto.
    cell(29, 9, "762", align=right)  # Ingresamos el texto.
    cell(29, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 30.
    cell(30, 2, "Facturas Activo Fijo", align=left_wrap)  # Ingresamos el texto.
    merge(30, 2, 30, 4)  # Juntamos las celdas.
    cell(30, 5, "", align=right)  # Ingresamos el texto.
    cell(30, 6, 0, align=right)  # Ingresamos el texto.
    cell(30, 7, "524", align=right)  # Ingresamos el texto.
    cell(30, 8, 0, align=right)  # Ingresamos el texto.
    cell(30, 9, "525", align=right)  # Ingresamos el texto.
    cell(30, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 31.
    cell(31, 2, "Notas de Crédito recibidos", align=left_wrap)  # Ingresamos el texto.
    merge(31, 2, 31, 4)  # Juntamos las celdas.
    cell(31, 5, "61", align=right)  # Ingresamos el texto.
    cell(31, 6, 0, align=right)  # Ingresamos el texto.
    cell(31, 7, "527", align=right)  # Ingresamos el texto.
    cell(31, 8, 0, align=right)  # Ingresamos el texto.
    cell(31, 9, "528", align=right)  # Ingresamos el texto.
    cell(31, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 32.
    cell(32, 2, "Notas de Débito recibidas", align=left_wrap)  # Ingresamos el texto.
    merge(32, 2, 32, 4)  # Juntamos las celdas.
    cell(32, 5, "56", align=right)  # Ingresamos el texto.
    cell(32, 6, 0, align=right)  # Ingresamos el texto.
    cell(32, 7, "531", align=right)  # Ingresamos el texto.
    cell(32, 8, 0, align=right)  # Ingresamos el texto.
    cell(32, 9, "532", align=right)  # Ingresamos el texto.
    cell(32, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 33.
    cell(33, 2, "Form. de pago de impotaciones del giro", align=left_wrap)  # Ingresamos el texto.
    merge(33, 2, 33, 4)  # Juntamos las celdas.
    cell(33, 5, "914", align=right)  # Ingresamos el texto.
    cell(33, 6, 0, align=right)  # Ingresamos el texto.
    cell(33, 7, "534", align=right)  # Ingresamos el texto.
    cell(33, 8, 0, align=right)  # Ingresamos el texto.
    cell(33, 9, "535", align=right)  # Ingresamos el texto.
    cell(33, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 34.
    cell(34, 2, "Form. de pago de impotaciones de activo fijo", align=left_wrap)  # Ingresamos el texto.
    merge(34, 2, 34, 4)  # Juntamos las celdas.
    cell(34, 5, "", align=right)  # Ingresamos el texto.
    cell(34, 6, 0, align=right)  # Ingresamos el texto.
    cell(34, 7, "536", align=right)  # Ingresamos el texto.
    cell(34, 8, 0, align=right)  # Ingresamos el texto.
    cell(34, 9, "553", align=right)  # Ingresamos el texto.
    cell(34, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 35.
    cell(35, 2, "Remanente Crédito Fiscal Mes Anterior UTM", align=left_wrap)  # Ingresamos el texto.
    merge(35, 2, 35, 4)  # Juntamos las celdas.
    merge(35, 5, 35, 8)  # Juntamos las celdas.
    cell(35, 9, "504", align=right)  # Ingresamos el texto.
    cell(35, 10, 0, align=right)  # Ingresamos el texto.

    # Linea 36.
    cell(36, 2, "")
    cell(36, 3, "TOTAL NETO", align=center, bold=True)  # Ingresamos el texto.
    merge(36, 3, 36, 5)  # Juntamos las celdas.
    cell(36, 6, 0, align=right, bold=True)  # Ingresamos el texto.
    cell(36, 7, "TOTAL IVA", align=center, bold=True)
    merge(36, 7, 36, 8)  # Juntamos las celdas.
    cell(36, 9, "537", align=center, bold=True)
    cell(36, 10, 0, align=right, bold=True)  # Ingresamos el texto.
    cell(36, 11, "Remanente", align=right, bold=True)  # Ingresamos el texto.

    # Linea 37.
    cell(37, 2, "IVA POR PAGAR", align=center, bold=True)  # Ingresamos el texto.
    merge(37, 2, 37, 6)  # Juntamos las celdas.
    cell(37, 7, "89", align=right)  # Ingresamos el texto.
    cell(37, 8, 0, align=right)  # Ingresamos el texto.
    merge(37, 8, 37, 10)  # Juntamos las celdas.
    cell(37, 11, 0, align=right)  # Ingresamos el texto.



    # ────────────────────────────────────────────────
    # RETENCIONES (remuneraciones y honorarios) (Es del 42 al 46 y del B al J horizontalmente).
    # Acá faltan datos y celdas por agregar, ojo.
    # ────────────────────────────────────────────────

    # Linea 38.
    cell(38, 2, "RETENCIONES", align=center, bold=True, font_size=11)
    merge(38, 2, 38, 10)

    # Linea .39
    cell(39, 2, "Impuesto Único a los trabajadores (Total Haberes)", align=left_wrap)  # Ingresamos el texto.
    merge(39, 2, 39, 6)  # Juntamos las celdas.
    cell(39, 7, "48", align=right)  # Ingresamos el texto.
    cell(39, 8, 0, align=right)  # Ingresamos el texto.
    cell(39, 9, 0, align=right)  # Ingresamos el texto.  HAY UN VALOR SIN CÓDIGO.
    merge(39, 9, 39, 10)  # Juntamos las celdas.

    # Linea 40.
    cell(40, 2, "Retención de Honorarios 14,5% Base (Honor)", align=left_wrap)  # Ingresamos el texto.
    merge(40, 2, 40, 6)  # Juntamos las celdas.
    cell(40, 7, "151", align=right)  # Ingresamos el texto.
    cell(40, 8, 0, align=right)  # Ingresamos el texto.
    cell(40, 9, 0, align=right)  # Ingresamos el texto.  HAY UN VALOR SIN CÓDIGO.
    merge(40, 9, 40, 10)  # Juntamos las celdas.

    # Linea 41.
    cell(41, 2, "Retenc.s/rtas.Art.42 N°2 LIR 21.252 con tasa 3% por reintegro (L.Remu)", align=left_wrap)  # Ingresamos el texto.
    merge(41, 2, 41, 6)  # Juntamos las celdas.
    cell(41, 7, "49", align=right)  # Ingresamos el texto.
    cell(41, 8, 0, align=right)  # Ingresamos el texto.
    cell(41, 9, 0, align=right)  # Ingresamos el texto.  HAY UN VALOR SIN CÓDIGO.
    merge(41, 9, 41, 10)  # Juntamos las celdas.

    # Linea 42.
    cell(42, 2, "Retenc.s/rtas.Art.42 N°2 LIR 21.242 con tasa (Honor)", align=left_wrap)  # Ingresamos el texto.
    merge(42, 2, 42, 6)  # Juntamos las celdas.
    cell(42, 7, "155", align=right)  # Ingresamos el texto.
    cell(42, 8, 0, align=right)  # Ingresamos el texto.
    cell(42, 9, 0, align=right)  # Ingresamos el texto.  HAY UN VALOR SIN CÓDIGO.
    merge(42, 9, 42, 10)  # Juntamos las celdas.



    # ────────────────────────────────────────────────
    # PPM (Es del 47 al 49 y del B a la I horizontalmente).
    # ────────────────────────────────────────────────

    # Linea 43.
    cell(43, 2, "PPM 1° CATEGORÍA", align=left_wrap)  # Ingresamos el texto.
    cell(43, 3, "BASE IMPONIBLE |115|", align=left_wrap)  # Ingresamos el texto.
    merge(43, 3, 43, 4)  # Juntamos las celdas.
    cell(43, 5, 0, align=left_wrap)  # Ingresamos el texto.
    merge(43, 5, 43, 6)  # Juntamos las celdas.
    cell(43, 7, "2%", align=left_wrap)  # Ingresamos el texto.
    merge(43, 7, 43, 8)  # Juntamos las celdas.
    cell(43, 9, 0, align=left_wrap)  # Ingresamos el texto.
    merge(43, 9, 43, 10)  # Juntamos las celdas.

    # Linea 44.
    cell(44, 2, "PPM 2° CATEGORÍA", align=left_wrap)  # Ingresamos el texto.
    cell(44, 3, "BASE", align=left_wrap)  # Ingresamos el texto.
    merge(44, 3, 44, 4)  # Juntamos las celdas.
    cell(44, 5, 0, align=left_wrap)  # Ingresamos el texto.
    merge(44, 5, 44, 6)  # Juntamos las celdas.
    cell(44, 7, 0, align=left_wrap)  # Ingresamos el texto.
    merge(44, 7, 44, 8)  # Juntamos las celdas.
    cell(44, 9, 0, align=left_wrap)  # Ingresamos el texto.
    merge(44, 9, 44, 10)  # Juntamos las celdas.

    # Linea 45.
    cell(45, 2, "PPM.TRANSPORTISTA", align=left_wrap)  # Ingresamos el texto.
    cell(45, 3, "BASE", align=left_wrap)  # Ingresamos el texto.
    merge(45, 3, 45, 4)  # Juntamos las celdas.
    cell(45, 5, 0, align=left_wrap)  # Ingresamos el texto.
    merge(45, 5, 45, 6)  # Juntamos las celdas.
    cell(45, 7, 0, align=left_wrap)  # Ingresamos el texto.
    merge(45, 7, 45, 8)  # Juntamos las celdas.
    cell(45, 9, 0, align=left_wrap)  # Ingresamos el texto.
    merge(45, 9, 45, 10)  # Juntamos las celdas.

    # Linea 46.
    merge(46, 2, 46, 6)  # Juntamos las celdas.
    cell(46, 7, "TOTAL", align=left_wrap, bold=True, font_size=11)  # Ingresamos el texto.
    merge(46, 7, 46, 8)  # Juntamos las celdas.
    cell(46, 9, 0, align=left_wrap, bold=True)  # Ingresamos el texto.
    merge(46, 9, 46, 10)  # Juntamos las celdas.

    # ────────────────────────────────────────────────
    # Confeccionado (Es del 51 al 52 y del B a la K horizontalmente)
    # ────────────────────────────────────────────────
    
    # Linea 47.
    cell(47, 2, "Confeccionado", align=left_wrap)  # Ingresamos el texto.
    cell(47, 3, "", align=left_wrap)  # Ingresamos el texto. ACÁ VA LA FECHA EN QUE SE HIZO ESTE DOCUMENTO.
    cell(47, 4, "A.V", align=center)  # Ingresamos el texto.
    merge(47, 4, 47, 6)  # Juntamos las celdas.

    # Ajustes de ancho para que se vean las casillas (el ajustar unas casillas parece esconder otras, ojo con eso).
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 6
    ws.column_dimensions['J'].width = 13
    ws.column_dimensions['K'].width = 13


    return wb