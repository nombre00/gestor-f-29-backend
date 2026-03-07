# Escritor que ingresa los datos a una plantilla.

# Bibliotecas.
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
# Módulos.
from f29_backend.domain.entities.resumenF29 import ResumenF29

# Flujo de acciones: recibimos el resumen, recibimos la plantilla resumen f-29, tomamos los datos del resumen, ingresamos esos datos a la plantilla del resumen f-29 y retornamos un workbook.

def resumenF29Escritor2(resumen: ResumenF29, plantilla: openpyxl.Workbook) -> openpyxl.Workbook:

    # Seleccionamos la hoja de la plantilla.
    ws = plantilla.active
    right = Alignment(horizontal='right')
    center = Alignment(horizontal='center')

    # Función helper para poner valor + borde + estilo
    # parámetros: r = fila; c = columna; font = fuente de la letra; alineacion; aplica el borde definido arriba.
    """ def cell(r, c, value="", font_name="Arial", font_size=10, bold=False, align=right):
        col_letter = openpyxl.utils.get_column_letter(c)  # .get_column_letter() convierte el numero de la columna a letra.
        ref = f"{col_letter}{r}"
        ws[ref] = value
        cell_obj = ws[ref]
        cell_obj.font = Font(name=font_name, size=font_size, bold=bold)
        if align:
            cell_obj.alignment = align """
    # Version que formatéa el número.
    def cell(r, c, value="", font_name="Arial", font_size=10, bold=False, 
         align=right, number_format=None):
        col_letter = openpyxl.utils.get_column_letter(c)
        ref = f"{col_letter}{r}"
        ws[ref] = value
        cell_obj = ws[ref]
        cell_obj.font = Font(name=font_name, size=font_size, bold=bold)
        if align:
            cell_obj.alignment = align
        if number_format is not None and isinstance(value, (int, float)):  # formateamos.
            cell_obj.number_format = number_format
        elif isinstance(value, (int, float)):  # default para montos
            cell_obj.number_format = '#,##0'      # entero → 1.234.567
    


    # Tomamos los datos del resumen y los asignamos a la plantilla de excel del resumen.
    # Falta formatear los datos (agregar puntos y comas).

    # Partimos con el encabezado.
    # Linea 3.
    mes = resumen.encabezado['periodo_mes']
    cell(3, 8, mes, bold=True, align=center, font_size=12)  # Ingresamos el texto.
    anio = resumen.encabezado['periodo_anio']
    cell(3, 10, anio, bold=True, align=center, font_size=12)  # Ingresamos el texto.

    # Linea 4.
    nombre = resumen.encabezado['nombre']
    cell(4, 3, nombre, align=center, font_size=11)

    # Linea 5.
    rut = resumen.encabezado['rut']
    cell(5, 3, rut, align=right, font_size=11) 
    folio = resumen.encabezado['numero']
    cell(5, 9, folio, bold=True, font_size=26, align=center)

    # Linea 6.
    clave = resumen.encabezado['clave_sii']
    cell(6, 3, clave, align=right, font_size=11)
    

    # seguimos por el detalle de las ventas (lista de diccionarios (cada diccionario una fila del excel)). 
    for fila in resumen.ventas_detalle:
        # Linea 9, Exportaciones.
        if fila.get('tipo') == 110:  # en esta fila ingresamos los valores de los códigos son = 110, 585 y 20.
            valor110 = fila.get('neto')  # código 110.
            cell(9, 6, valor110 , align=right)  # Ingresamos el texto.
            valor585 = fila.get('td')  # código 585.
            cell(9, 8, valor585, align=right)  # Ingresamos el texto.
            valor20 = fila.get('iva')  # código 20.
            cell(9, 10, valor20, align=right)  # Ingresamos el texto.

        # Linea 10, Ventas y/o servicios no gravados.
        if fila.get('tipo') == 34:  # en esta fila ingresamos los valores de los códigos son = 34, 586 y 142.
            valor34 = fila.get('neto')  # código 34.
            cell(10, 6, valor34 , align=right)  # Ingresamos el texto.
            valor586 = fila.get('td')  # código 586.
            cell(10, 8, valor586, align=right)  # Ingresamos el texto.
            valor142 = fila.get('iva')  # código 142.
            cell(10, 10, valor142, align=right)  # Ingresamos el texto.

        # Linea 11, Facturas de compra recibidas con retención total.
        if fila.get('tipo') == 46:  # en esta fila ingresamos los valores de los códigos son = 46, 515 y 587.
            valor46 = fila.get('neto')  # código 46.
            cell(11, 6, valor46, align=right)  # Ingresamos el texto.
            valor515 = fila.get('td')  # código 515.
            cell(11, 8, valor515, align=right)  # Ingresamos el texto.
            valor587 = fila.get('iva')  # código 587.
            cell(11, 10, valor587, align=right)  # Ingresamos el texto.

        # Linea 12, facturas emitidas.
        if fila.get('tipo') == 33:  # en esta fila ingresamos los valores de los códigos son = 33, 503 y 502.
            valor33 = fila.get('neto')  # código 33.
            cell(12, 6, valor33 , align=right)  # Ingresamos el texto.
            valor503 = fila.get('td')  # código 503.
            cell(12, 8, valor503, align=right)  # Ingresamos el texto.
            valor502 = fila.get('iva')  # código 502.
            cell(12, 10, valor502, align=right)  # Ingresamos el texto.
        
        # Linea 13, Cant.de Doctos.Boletas.
        if fila.get('tipo') == 39:  # en esta fila ingresamos los valores de los códigos son = 39, 110 y 111.
            valor39 = fila.get('neto')  # código 39.
            cell(13, 6, valor39 , align=right)  # Ingresamos el texto.
            valor110 = fila.get('td')  # código 110.
            cell(13, 8, valor110, align=right)  # Ingresamos el texto.
            valor111 = fila.get('iva')  # código 111.
            cell(13, 10, valor111, align=right)  # Ingresamos el texto.

        # Linea 14, Comprob.o Recibod Trasacciones Medios.
        if fila.get('tipo') == 48:  # en esta fila ingresamos los valores de los códigos son = 48, 758, 759.
            valor48 = fila.get('neto')  # código 48.
            cell(14, 6, valor48 , align=right)  # Ingresamos el texto.
            valor758 = fila.get('td')  # código 758.
            cell(14, 8, valor758, align=right)  # Ingresamos el texto.
            valor759 = fila.get('iva')  # código 759.
            cell(14, 10, valor759, align=right)  # Ingresamos el texto.
        
        # Linea 15, Notas de Débito emitidas.
        if fila.get('tipo') == 56:  # en esta fila ingresamos los valores de los códigos son = 56, 512, 513.
            valor56 = fila.get('neto')  # código 56.
            cell(15, 6, valor56 , align=right)  # Ingresamos el texto.
            valor512 = fila.get('td')  # código 512.
            cell(15, 8, valor512, align=right)  # Ingresamos el texto.
            valor513 = fila.get('iva')  # código 513.
            cell(15, 10, valor513, align=right)  # Ingresamos el texto.
        
        # Linea 16, Notas de Crédito emitidas.
        if fila.get('tipo') == 61:  # en esta fila ingresamos los valores de los códigos son = 61, 509, 510.
            valor61 = fila.get('neto')  # código 61.
            cell(16, 6, valor61 , align=right)  # Ingresamos el texto.
            valor509 = fila.get('td')  # código 509.
            cell(16, 8, valor509, align=right)  # Ingresamos el texto.
            valor510 = fila.get('iva')  # código 510.
            cell(16, 10, valor510, align=right)  # Ingresamos el texto.
        
        # Linea 17, Facturas de Compras recibidas.
        if fila.get('tipo') == 45:  # en esta fila ingresamos los valores de los códigos son = 45, 516, 517.
            valor45 = fila.get('neto')  # código 45.
            cell(17, 6, valor45 , align=right)  # Ingresamos el texto.
            valor516 = fila.get('td')  # código 516.
            cell(17, 8, valor516, align=right)  # Ingresamos el texto.
            valor517 = fila.get('iva')  # código 517.
            cell(17, 10, valor517, align=right)  # Ingresamos el texto.

        # Linea 18, Liquidación y Liquidación Factura.
        if fila.get('tipo') == 43:  # en esta fila ingresamos los valores de los códigos son = 43, 500, 501.
            valor43 = fila.get('neto')  # código 43.
            cell(18, 6, valor43 , align=right)  # Ingresamos el texto.
            valor500 = fila.get('td')  # código 500.
            cell(18, 8, valor500, align=right)  # Ingresamos el texto.
            valor501 = fila.get('iva')  # código 510.
            cell(18, 10, valor501, align=right)  # Ingresamos el texto. 


    # Seguimos con ventas_total.
    # Linea 19 (tiene que ser un cálculo).
    ventasTotal = resumen.ventas_total.get('neto')   # en esta fila ingresamos los valores de los códigos son = 538.
    cell(19, 6, ventasTotal, align=right, bold=True)  # Ingresamos el texto.
    ventasIvaTotal = resumen.ventas_total.get('iva')
    cell(19, 10, ventasIvaTotal, align=right, bold=True)  # Ingresamos el texto.


    # Seguimos por el detalle de las compras (lista de diccionarios (cada diccionario una fila del excel)).
    # Linea 22 (ese valor se toma del código 89), no es necesario iterar.
    valor511 = resumen.IVAPP  # código 511.
    cell(22, 6, valor511, align=right)  # Ingresamos el texto.
    ##### Preguntar de donde saco el valor del codigo 514 de esta misma linea. #####
     
    for fila in resumen.compras_detalle:
        # Linea 24, Compras Internas Afectadas (Sin derecho a Cdto.Fiscal).
        ### Preguntar de donde saco los valores que acá se ingresan ###

        # Linea 25, Importaciones (Sin derecho a Cdto.Fiscal).
        ### Preguntar de donde saco los valores que acá se ingresan ###

        # Linea 26, Compras Internas exentas o No gravadas.
        if fila.get('tipo') == 34:  # en esta fila ingresamos los valores de los códigos son = 34, 584, 562.
            valor34 = fila.get('neto')  # código 34.
            cell(28, 6, valor34 , align=right)  # Ingresamos el texto.
            valor584 = fila.get('td')  # código 584.
            cell(28, 8, valor584, align=right)  # Ingresamos el texto.
            valor562 = fila.get('iva_rec')  # código 562.
            cell(28, 10, valor562, align=right)  # Ingresamos el texto.

        # Linea 28, Facturas recibidas y Facturas de compra.
        if fila.get('tipo') == 33:  # en esta fila ingresamos los valores de los códigos son = 33, 519 y 520.
            valor33 = fila.get('neto')  # código 33.
            cell(28, 6, valor33 , align=right)  # Ingresamos el texto.
            valor519 = fila.get('td')  # código 519.
            cell(28, 8, valor519, align=right)  # Ingresamos el texto.
            valor520 = fila.get('iva_rec')  # código 520.
            cell(28, 10, valor520, align=right)  # Ingresamos el texto.
        
        # Linea 29, Fact.Recib.Prov.Superm.Art.33 N°4 DL825(Ley 20,780).
        if fila.get('tipo') == 761:
            valorsuper = fila.get('neto')  # valorsuper.
            cell(29, 6, valorsuper , align=right)  # Ingresamos el texto.
            valor761 = fila.get('td')  # código 761.
            cell(29, 8, valor761, align=right)  # Ingresamos el texto.
            valor762 = fila.get('iva_rec')  # código 762.
            cell(29, 10, valor762, align=right)  # Ingresamos el texto.

        # Linea 30, Facturas Activo Fijo.
        if fila.get('tipo') == 524:
            valorActivoFijo = fila.get('neto')  # código 524.
            cell(30, 6, valorActivoFijo , align=right)  # Ingresamos el texto.
            valor524 = fila.get('td')  # código 524.
            cell(30, 8, valor524, align=right)  # Ingresamos el texto.
            valor525 = fila.get('iva_rec')  # código 525.
            cell(30, 10, valor525, align=right)  # Ingresamos el texto.
        
        # Linea 31, Notas de Crédito recibidos.
        if fila.get('tipo') == 61:  # en esta fila ingresamos los valores de los códigos son = 61, 527 y 528.
            valor61 = fila.get('neto')  # código 61.
            cell(31, 6, valor61 , align=right)  # Ingresamos el texto.
            valor527 = fila.get('td')  # código 527.
            cell(31, 8, valor527, align=right)  # Ingresamos el texto.
            valor528 = fila.get('iva_rec')  # código 528.
            cell(31, 10, valor528, align=right)  # Ingresamos el texto.

        # Linea 32, Notas de Débito recibidas.
        if fila.get('tipo') == 56:  # en esta fila ingresamos los valores de los códigos son = 56, 531, 532
            valor56 = fila.get('neto')  # código 56.
            cell(32, 6, valor61 , align=right)  # Ingresamos el texto.
            valor531 = fila.get('td')  # código 531.
            cell(32, 8, valor531, align=right)  # Ingresamos el texto.
            valor532 = fila.get('iva_rec')  # código 532.
            cell(32, 10, valor532, align=right)  # Ingresamos el texto.

        # Linea 33, Form.depago de impotaciones del giro.
        if fila.get('tipo') == 914:  # en esta fila ingresamos los valores de los códigos son = 914, 534, 534
            valor914 = fila.get('neto')  # código 914.
            cell(33, 6, valor914 , align=right)  # Ingresamos el texto.
            valor534 = fila.get('td')  # código 545.
            cell(33, 8, valor534, align=right)  # Ingresamos el texto.
            valor535 = fila.get('iva_rec')  # código 535.
            cell(33, 10, valor535, align=right)  # Ingresamos el texto.

        # Linea 34,Form.depago deimpotaciones de activo fijo.
        if fila.get('tipo') == 536:  # en esta fila ingresamos los valores de los códigos son = 914, 534, 534
            codActivoFijoImportacion = fila.get('neto')  # codActivoFijoImportacion
            cell(34, 6, codActivoFijoImportacion , align=right)  # Ingresamos el texto.
            valor536 = fila.get('td')  # código 536.
            cell(34, 8, valor536, align=right)  # Ingresamos el texto.
            valor553 = fila.get('iva_rec')  # código 553.
            cell(34, 10, valor553, align=right)  # Ingresamos el texto.

        # Linea 35, Remanente Crédito Fiscal Mes Anterior UTM.
        remanenteMesAnterior = resumen.remanenteMesAnterior 
        cell(35, 10, remanenteMesAnterior , align=right)  # Ingresamos el texto.
    

    # Continuamos en el total de las compras.
    # Linea 36 esto (tiene que ser un cálculo).
    comprasTotal = resumen.compras_total.get('neto')   # en esta fila ingresamos los valores de los códigos son = 537.
    cell(36, 6, comprasTotal, align=right, bold=True)  # Ingresamos el texto.
    comprasIvaTotal = resumen.compras_total.get('iva_rec')
    cell(36, 10, comprasIvaTotal, align=right, bold=True)  # Ingresamos el texto.


    # Continuamos con IVA por pagar.
    # Linea 37. 
    iva_por_pagar = resumen.IVAPP
    remanente = resumen.remanente
    cell(37, 8, iva_por_pagar, align=right, bold=True)  # Ingresamos el texto.
    cell(37, 11, remanente, align=right, bold=True)  # Ingresamos el texto.


    # Continuamos con remuneraciones y honorarios.
    # Linea 39, Impuesto Único a los trabajadores (Total Haberes).
    totalHaberes = resumen.remuneraciones.get('th_remuneraciones')
    cell(39, 8, totalHaberes, align=right)  # Ingresamos el texto.
    impuestoUnico = resumen.remuneraciones.get('impt_unico')
    cell(39, 9, impuestoUnico, align=right)  # Ingresamos el texto.

    # Linea 40, Retención de Honorarios 14,5% Base (Honor). 
    honorarios = resumen.honorarios.get('honorarios')
    cell(40, 8, honorarios, align=right)  # Ingresamos el texto.
    retencion = resumen.honorarios.get('retencion')
    cell(40, 9, retencion, align=right)  # Ingresamos el texto.

    # Linea 41, Retenc.s/rtas.Art.42 N°2 LIR 21.252 con tasa 3% por reintegro (L.Remu).
    totalHaberes = resumen.remuneraciones.get('th_remuneraciones')
    cell(41, 8, totalHaberes, align=right)  # Ingresamos el texto.
    remuneraciones3porciento = resumen.remuneraciones.get('rem_3porc')
    cell(41, 9, remuneraciones3porciento, align=right)  # Ingresamos el texto.

    # Linea 42, Retenc.s/rtas.Art.42 N°2 LIR 21.242 con tasa (Honor). 
    cod155 = resumen.honorarios.get('cod155')  # Ingresamos el texto.
    cell(41, 9, cod155, align=right)  # Ingresamos el texto.


    # Continuamos con PPM.
    # Linea 43, PPM 1° CATEGORÍA.
    baseImponible = resumen.ppm.get('base')
    cell(43, 5, baseImponible, align=right)  # Ingresamos el texto.
    tasa = resumen.ppm.get('tasa')
    cell(43, 7, tasa, align=right)  # Ingresamos el texto.
    ppm = resumen.ppm.get('ppm')
    cell(43, 9, ppm, align=right)  # Ingresamos el texto.

    # Linea 44, PPM 2° CATEGORÍA.
    ppm2_base = resumen.ppm.get('PPM2_base')
    cell(44, 5, ppm2_base, align=right)  # Ingresamos el texto.
    ppm2_tasa = resumen.ppm.get('PPM2_tasa')
    cell(44, 7, ppm2_tasa, align=right)  # Ingresamos el texto.
    ppm2_valor = resumen.ppm.get('PPM2_valor')
    cell(44, 9, ppm2_valor, align=right)  # Ingresamos el texto.

    # Linea 45, PPM.TRANSPORTISTA.
    ppm_transportista_base = resumen.ppm.get('PPM_transportista_base')
    cell(45, 5, ppm_transportista_base, align=right)  # Ingresamos el texto.
    ppm_transportista_tasa = resumen.ppm.get('PPM_transportista_tasa')
    cell(45, 7, ppm_transportista_tasa, align=right)  # Ingresamos el texto.
    ppm_transportista_valor = resumen.ppm.get('PPM_transportista_valor')
    cell(45, 9, ppm_transportista_valor, align=right)  # Ingresamos el texto.

    # Linea 46 (total final) (tiene que ser un cálculo).
    tt = resumen.TT
    cell(46, 9, tt, align=right)  # Ingresamos el texto.


    # Terminamos con los datos extra.
    # Linea 49.
    arriendos_pagados = resumen.arriendos_pagados
    cell(49, 5, arriendos_pagados, align=right)
    # Linea 50.
    gastos_generales_boletas = resumen.gastos_generales_boletas
    cell(50, 5, gastos_generales_boletas, align=right)


    # Imprimimos en pantalla la plantilla.
    for fila in plantilla:
        print(fila)


    # Retornamos el excel.
    return plantilla